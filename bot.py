import logging
import os
import re
import threading
import time
import requests
from itertools import product
from telegram import Update, Bot
from telegram.ext import Application, CommandHandler, ContextTypes
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
ODDS_API_KEY = os.getenv("ODDS_API_KEY")
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
MISE_DEFAUT = 100000
TAXE = 0.10
SEUIL_BENEFICE_FCFA = 500   # +500 FCFA benefice net minimum pour ✅
SEUIL_ALERTE = 0.10
INTERVALLE_SCAN = 1800
ODDS_API_URL = "https://api.the-odds-api.com/v4"
historique = {}
bankroll_data = {}
objectif_data = {}
abonnes_alertes = set()
alertes_envoyees = set()
# ─────────────────────────────────────────────
# BOOKMAKERS EXCLUS (suspendus, en litige fiscal, etc.)
# Liste par defaut modifiable via la variable Railway BOOKMAKERS_EXCLUS
# (noms separes par des virgules). Modifiable a chaud via /exclure et /inclure,
# mais reinitialisee a la valeur par defaut/env a chaque redemarrage du bot.
# ─────────────────────────────────────────────
BOOKMAKERS_EXCLUS_DEFAUT = "1xBet,Premier Bet,Bet223"
bookmakers_exclus = set(
    b.strip().lower()
    for b in os.getenv("BOOKMAKERS_EXCLUS", BOOKMAKERS_EXCLUS_DEFAUT).split(",")
    if b.strip()
)
# ─────────────────────────────────────────────
# MOTEUR DE CALCUL
# ─────────────────────────────────────────────
class AlboraEngine:
    @staticmethod
    def detect_arbitrage(cotes):
        s = sum(1 / c for c in cotes)
        if s < 1:
            profit = round((1 / s - 1) * 100, 2)
            stakes = {f"cote_{i+1}": round((1 / c / s) * 100, 2) for i, c in enumerate(cotes)}
            return {"arbitrage": True, "profit": profit, "stakes": stakes, "sum": round(s, 4)}
        return {"arbitrage": False, "sum": round(s, 4)}
    @staticmethod
    def combo_matchs(matchs, noms_matchs, mise_totale):
        labels = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        resultats = []
        nb_total = 1
        for m in matchs:
            nb_total *= len(m)
        mise_par_combo = round(mise_totale / nb_total, 0)
        for combo in product(*[range(len(m)) for m in matchs]):
            cotes_sel = [matchs[i][combo[i]] for i in range(len(matchs))]
            noms_combo = []
            for i, idx in enumerate(combo):
                match_nom = noms_matchs[i] if i < len(noms_matchs) else f"Match {i+1}"
                equipes = match_nom.split(" vs ")
                if len(equipes) == 2:
                    if idx == 0:
                        noms_combo.append(equipes[0].strip())
                    elif idx == 1:
                        noms_combo.append("Nul")
                    else:
                        noms_combo.append(equipes[1].strip())
                else:
                    noms_combo.append(f"{match_nom}({idx+1})")
            cote_totale = 1.0
            for c in cotes_sel:
                cote_totale *= c
            cote_totale = round(cote_totale, 4)
            gain_brut = round(mise_par_combo * cote_totale, 0)
            taxe = round(gain_brut * TAXE, 0)
            gain_net = round(gain_brut - taxe, 0)
            benefice_net = round(gain_net - mise_totale, 0)
            benefice_pct = round((gain_net / mise_totale - 1) * 100, 2)
            prob = round((1 / cote_totale) * 100, 2)
            rentable = benefice_net >= SEUIL_BENEFICE_FCFA
            resultats.append({
                "noms_combo": " + ".join(noms_combo),
                "cote_totale": cote_totale,
                "mise_par_combo": int(mise_par_combo),
                "gain_brut": int(gain_brut),
                "taxe": int(taxe),
                "gain_net": int(gain_net),
                "benefice_net": int(benefice_net),
                "benefice_pct": benefice_pct,
                "prob": prob,
                "rentable": rentable,
            })
        resultats.sort(key=lambda x: x["cote_totale"])
        return resultats, nb_total, int(mise_par_combo)
    @staticmethod
    def detect_value(cotes):
        total_prob = sum(1 / c for c in cotes)
        marge = round((total_prob - 1) * 100, 2)
        values = []
        for i, c in enumerate(cotes):
            prob_implicite = round((1 / c) / total_prob * 100, 2)
            prob_juste = round(1 / c * 100, 2)
            edge = round(prob_juste - prob_implicite, 2)
            cote_juste = round(1 / (prob_implicite / 100), 3)
            values.append({
                "index": i + 1,
                "cote": c,
                "prob_implicite": prob_implicite,
                "prob_juste": prob_juste,
                "cote_juste": cote_juste,
                "edge": edge,
                "value": c > cote_juste,
            })
        return values, marge
    @staticmethod
    def evaluer_risque(cotes_combo):
        cote_totale = 1.0
        for c in cotes_combo:
            cote_totale *= c
        prob = round((1 / cote_totale) * 100, 2)
        if prob >= 40:
            niveau, emoji = "FAIBLE", "🟢"
        elif prob >= 20:
            niveau, emoji = "MOYEN", "🟡"
        elif prob >= 10:
            niveau, emoji = "ELEVE", "🟠"
        else:
            niveau, emoji = "TRES ELEVE", "🔴"
        return {
            "cote_totale": round(cote_totale, 4),
            "prob": prob,
            "niveau": niveau,
            "emoji": emoji,
            "nb_matchs": len(cotes_combo),
        }
# ─────────────────────────────────────────────
# ODDS API
# ─────────────────────────────────────────────
class OddsEngine:
    @staticmethod
    def get_sports():
        if not ODDS_API_KEY:
            return []
        try:
            r = requests.get(f"{ODDS_API_URL}/sports", params={"apiKey": ODDS_API_KEY}, timeout=10)
            if r.status_code == 200:
                return r.json()
        except Exception as e:
            logging.error(f"get_sports error: {e}")
        return []
    @staticmethod
    def get_odds(sport_key):
        if not ODDS_API_KEY:
            return []
        try:
            r = requests.get(
                f"{ODDS_API_URL}/sports/{sport_key}/odds",
                params={"apiKey": ODDS_API_KEY, "regions": "eu", "markets": "h2h", "oddsFormat": "decimal"},
                timeout=15
            )
            if r.status_code == 200:
                return r.json()
        except Exception as e:
            logging.error(f"get_odds error: {e}")
        return []
    @staticmethod
    def analyser_matchs(events):
        opportunites = []
        for event in events:
            home = event.get("home_team", "?")
            away = event.get("away_team", "?")
            commence = event.get("commence_time", "")
            bookmakers = event.get("bookmakers", [])
            if not bookmakers:
                continue
            best = {"home": 0, "draw": 0, "away": 0, "bk_home": "", "bk_draw": "", "bk_away": ""}
            for bk in bookmakers:
                bk_name = bk.get("title", "?")
                if bk_name.strip().lower() in bookmakers_exclus:
                    continue
                for market in bk.get("markets", []):
                    if market.get("key") != "h2h":
                        continue
                    for outcome in market.get("outcomes", []):
                        name = outcome.get("name", "")
                        price = outcome.get("price", 0)
                        if name == home and price > best["home"]:
                            best["home"] = price
                            best["bk_home"] = bk_name
                        elif name == away and price > best["away"]:
                            best["away"] = price
                            best["bk_away"] = bk_name
                        elif name == "Draw" and price > best["draw"]:
                            best["draw"] = price
                            best["bk_draw"] = bk_name
            cotes = [c for c in [best["home"], best["draw"], best["away"]] if c > 1]
            if len(cotes) < 2:
                continue
            arb = AlboraEngine.detect_arbitrage(cotes)
            values, marge = AlboraEngine.detect_value(cotes)
            gain_brut = round(MISE_DEFAUT * max(cotes), 0)
            taxe = round(gain_brut * TAXE, 0)
            gain_net = round(gain_brut - taxe, 0)
            benefice_net = round(gain_net - MISE_DEFAUT, 0)
            benefice_pct = round((gain_net / MISE_DEFAUT - 1) * 100, 2)
            opportunites.append({
                "match": f"{home} vs {away}",
                "heure": commence[:16].replace("T", " ") if commence else "?",
                "cotes": cotes,
                "best": best,
                "arbitrage": arb,
                "value_cotes": [v for v in values if v["value"]],
                "marge": marge,
                "gain_net": gain_net,
                "benefice_net": benefice_net,
                "benefice_pct": benefice_pct,
            })
        return opportunites
# ─────────────────────────────────────────────
# SCANNER AUTOMATIQUE
# ─────────────────────────────────────────────
def scanner_loop(bot_token: str):
    bot = Bot(token=bot_token)
    logging.info("Scanner automatique demarre.")
    while True:
        try:
            if abonnes_alertes and ODDS_API_KEY:
                logging.info("Scan en cours...")
                sports = OddsEngine.get_sports()
                alertes_session = []
                for sport in sports:
                    if not sport.get("active"):
                        continue
                    events = OddsEngine.get_odds(sport.get("key", ""))
                    if not events:
                        continue
                    for opp in OddsEngine.analyser_matchs(events):
                        match_id = f"{opp['match']}_{opp['heure']}"
                        if opp["arbitrage"]["arbitrage"] and match_id + "_arb" not in alertes_envoyees:
                            alertes_envoyees.add(match_id + "_arb")
                            alertes_session.append({"type": "ARBITRAGE", "sport": sport.get("title", ""), "opp": opp})
                        elif opp["benefice_net"] >= SEUIL_BENEFICE_FCFA and match_id + "_val" not in alertes_envoyees:
                            alertes_envoyees.add(match_id + "_val")
                            alertes_session.append({"type": "VALUE", "sport": sport.get("title", ""), "opp": opp})
                    time.sleep(0.5)
                for alerte in alertes_session:
                    msg = _formater_alerte(alerte)
                    for uid in list(abonnes_alertes):
                        try:
                            import asyncio
                            asyncio.run(bot.send_message(chat_id=uid, text=msg))
                        except Exception as e:
                            logging.error(f"Envoi alerte: {e}")
                logging.info(f"Scan termine. {len(alertes_session)} alerte(s).")
        except Exception as e:
            logging.error(f"Scanner error: {e}")
        time.sleep(INTERVALLE_SCAN)
def _formater_alerte(alerte):
    opp = alerte["opp"]
    best = opp["best"]
    sport = alerte["sport"]
    if alerte["type"] == "ARBITRAGE":
        arb = opp["arbitrage"]
        return (
            f"ARBITRAGE DETECTE\n{'='*30}\n"
            f"Sport: {sport}\nMatch: {opp['match']}\nHeure: {opp['heure']}\n\n"
            f"Profit garanti: +{arb['profit']}%\nSomme: {arb['sum']}\n\n"
            f"Meilleures cotes:\n"
            f"1 ({best['bk_home']}): {best['home']}\n"
            f"N ({best['bk_draw']}): {best['draw']}\n"
            f"2 ({best['bk_away']}): {best['away']}\n\n"
            f"Misez maintenant sur 1xBet."
        )
    else:
        signe = "+" if opp["benefice_pct"] >= 0 else ""
        return (
            f"ALERTE OPPORTUNITE\n{'='*30}\n"
            f"Sport: {sport}\nMatch: {opp['match']}\nHeure: {opp['heure']}\n\n"
            f"Meilleures cotes:\n"
            f"1 ({best['bk_home']}): {best['home']}\n"
            f"N ({best['bk_draw']}): {best['draw']}\n"
            f"2 ({best['bk_away']}): {best['away']}\n\n"
            f"Simulation {MISE_DEFAUT:,} FCFA:\n"
            f"Gain brut: {int(opp['gain_net'] / 0.9):,} FCFA\n"
            f"Taxe (10%): -{int(opp['gain_net'] / 0.9 * 0.1):,} FCFA\n"
            f"Gain net: {int(opp['gain_net']):,} FCFA ({signe}{opp['benefice_pct']}%)\n"
            f"Benefice net: +{int(opp['benefice_net']):,} FCFA\n\n"
            f"Misez maintenant sur 1xBet."
        )
# ─────────────────────────────────────────────
# COMMANDES TELEGRAM
# ─────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = (
        "ALBORAA BOT - Arbitrage et Combinaisons\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "ANALYSE MANUELLE\n"
        '/combo "Barça vs Real" 1.30 2.21 3.12 | "Mali vs Sénégal" 1.45 3.10\n'
        "/arbitrage 1.21 3.02 3.10\n"
        "/value 1.30 2.21 3.12\n"
        "/risque 1.30 2.21 | 1.45 3.10\n"
        "/simuler 1.85 10000\n\n"
        "ALERTES AUTOMATIQUES\n"
        "/alertes_on — Activer les alertes\n"
        "/alertes_off — Desactiver les alertes\n"
        "/scan — Scan manuel immediat\n"
        "/sports — Sports surveilles\n"
        "/exclure NomBookmaker — Exclure un bookmaker des scans\n"
        "/inclure NomBookmaker — Reactiver un bookmaker\n"
        "/liste_exclus — Voir les bookmakers exclus\n\n"
        "GESTION\n"
        "/bankroll 500000\n"
        "/bankroll_pari 5000 15000\n"
        "/bankroll_bilan\n"
        "/objectif 1000000\n"
        "/objectif_bilan\n\n"
        "HISTORIQUE\n"
        "/bilan | /effacer | /export\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"Devise: Franc CFA\n"
        f"Mise defaut: {MISE_DEFAUT:,} FCFA\n"
        f"Taxe Mali: {int(TAXE*100)}%\n"
        f"Seuil rentabilite: +{SEUIL_BENEFICE_FCFA:,} FCFA benefice net\n"
        f"Scan: toutes les 30 minutes"
    )
    await update.message.reply_text(msg)
async def alertes_on(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ODDS_API_KEY:
        await update.message.reply_text(
            "ODDS_API_KEY manquante.\n"
            "Ajoutez-la dans Railway > Variables > ODDS_API_KEY\n"
            "Cle gratuite sur : the-odds-api.com"
        )
        return
    abonnes_alertes.add(uid)
    await update.message.reply_text(
        f"Alertes activees.\n"
        f"Scan toutes les 30 minutes.\n"
        f"Seuil : +{SEUIL_BENEFICE_FCFA:,} FCFA benefice net minimum.\n"
        f"Utilisez /alertes_off pour desactiver."
    )
async def alertes_off(update: Update, context: ContextTypes.DEFAULT_TYPE):
    abonnes_alertes.discard(update.effective_user.id)
    await update.message.reply_text("Alertes desactivees.")
async def scan_manuel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not ODDS_API_KEY:
        await update.message.reply_text("ODDS_API_KEY manquante. Cle gratuite sur : the-odds-api.com")
        return
    await update.message.reply_text("Scan en cours... patientez.")
    try:
        sports = OddsEngine.get_sports()
        total_alertes = []
        for sport in sports[:10]:
            if not sport.get("active"):
                continue
            events = OddsEngine.get_odds(sport.get("key", ""))
            if not events:
                continue
            for opp in OddsEngine.analyser_matchs(events):
                if opp["arbitrage"]["arbitrage"] or opp["benefice_net"] >= SEUIL_BENEFICE_FCFA:
                    total_alertes.append({
                        "type": "ARBITRAGE" if opp["arbitrage"]["arbitrage"] else "VALUE",
                        "sport": sport.get("title", ""),
                        "opp": opp
                    })
            time.sleep(0.3)
        if not total_alertes:
            await update.message.reply_text(f"Scan termine. Aucune opportunite detectee.")
        else:
            await update.message.reply_text(f"{len(total_alertes)} opportunite(s) trouvee(s):")
            for alerte in total_alertes[:5]:
                await update.message.reply_text(_formater_alerte(alerte))
    except Exception as e:
        await update.message.reply_text(f"Erreur scan: {str(e)[:200]}")
async def exclure_bookmaker(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text(
            "Format: /exclure NomBookmaker\nExemple: /exclure Melbet"
        )
        return
    nom = " ".join(context.args).strip().lower()
    bookmakers_exclus.add(nom)
    liste = ", ".join(sorted(bookmakers_exclus)) or "aucune"
    await update.message.reply_text(
        f"'{nom}' exclu des scans.\nListe actuelle: {liste}\n\n"
        f"Note: reinitialisee a la liste par defaut si le bot redemarre."
    )
async def inclure_bookmaker(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text(
            "Format: /inclure NomBookmaker\nExemple: /inclure 1xBet"
        )
        return
    nom = " ".join(context.args).strip().lower()
    bookmakers_exclus.discard(nom)
    liste = ", ".join(sorted(bookmakers_exclus)) or "aucune"
    await update.message.reply_text(f"'{nom}' reactive.\nListe actuelle: {liste}")
async def liste_exclus(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not bookmakers_exclus:
        await update.message.reply_text("Aucun bookmaker exclu actuellement.")
        return
    lignes = ["BOOKMAKERS EXCLUS DES SCANS"] + [f"- {b}" for b in sorted(bookmakers_exclus)]
    await update.message.reply_text("\n".join(lignes))
async def sports_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not ODDS_API_KEY:
        await update.message.reply_text("ODDS_API_KEY manquante. Cle gratuite sur : the-odds-api.com")
        return
    try:
        sport_list = OddsEngine.get_sports()
        actifs = [s for s in sport_list if s.get("active")]
        lignes = [f"Sports surveilles : {len(actifs)}\n"]
        for s in actifs[:20]:
            lignes.append(f"- {s.get('title', s.get('key'))}")
        if len(actifs) > 20:
            lignes.append(f"... et {len(actifs)-20} autres.")
        await update.message.reply_text("\n".join(lignes))
    except Exception as e:
        await update.message.reply_text(f"Erreur: {e}")
async def arbitrage(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if len(args) < 2 or len(args) > 3:
        await update.message.reply_text("Format: /arbitrage 1.21 3.02 3.10")
        return
    try:
        cotes = list(map(float, args))
        result = AlboraEngine.detect_arbitrage(cotes)
        if result["arbitrage"]:
            stakes_str = "\n".join([f"Cote {k[-1]} ({cotes[int(k[-1])-1]}) -> {v}%" for k, v in result["stakes"].items()])
            msg = f"ARBITRAGE DETECTE\nProfit: {result['profit']}%\nSomme: {result['sum']}\n\nRepartition:\n{stakes_str}"
            _save_historique(update.effective_user.id, "arbitrage", cotes, result)
        else:
            msg = f"Pas d'arbitrage. Somme: {result['sum']} (doit etre < 1.00)"
        await update.message.reply_text(msg)
    except ValueError:
        await update.message.reply_text("Cotes invalides. Exemple: /arbitrage 1.21 3.02 3.10")
async def value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args or len(context.args) < 2:
        await update.message.reply_text("Format: /value 1.30 2.21 3.12")
        return
    try:
        cotes = list(map(float, context.args))
        values, marge = AlboraEngine.detect_value(cotes)
        lignes = []
        for v in values:
            emoji = "✅" if v["value"] else "❌"
            tag = "VALUE" if v["value"] else "NON VALUE"
            lignes.append(
                f"{emoji} Cote {v['index']}: {v['cote']}\n"
                f"   Prob marche: {v['prob_implicite']}% | Prob juste: {v['prob_juste']}%\n"
                f"   Cote juste: {v['cote_juste']} | Edge: {v['edge']:+}%\n"
                f"   -> {tag}"
            )
        msg = (
            f"ANALYSE VALUE\nMarge bookmaker: {marge}%\n{'='*30}\n\n"
            + "\n\n".join(lignes)
            + f"\n\n{'='*30}\nCote VALUE = bookmaker sous-estime la probabilite."
        )
        await update.message.reply_text(msg)
    except ValueError:
        await update.message.reply_text("Cotes invalides. Exemple: /value 1.30 2.21 3.12")
async def risque(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Format: /risque 1.30 2.21 | 1.45 3.10")
        return
    try:
        ligne = " ".join(context.args)
        cotes_combo = []
        for seg in ligne.split("|"):
            cotes_combo.extend([float(x) for x in seg.strip().split() if x.strip()])
        r = AlboraEngine.evaluer_risque(cotes_combo)
        gain_brut = round(MISE_DEFAUT * r["cote_totale"], 0)
        taxe = round(gain_brut * TAXE, 0)
        gain_net = round(gain_brut - taxe, 0)
        msg = (
            f"EVALUATION DU RISQUE\n{'='*30}\n"
            f"Matchs combines: {r['nb_matchs']}\n"
            f"Cote totale: {r['cote_totale']}\n"
            f"Probabilite: {r['prob']}%\n\n"
            f"{r['emoji']} Niveau: {r['niveau']}\n\n"
            f"Simulation {MISE_DEFAUT:,} FCFA:\n"
            f"Gain brut: {int(gain_brut):,} FCFA\n"
            f"Taxe (10%): -{int(taxe):,} FCFA\n"
            f"Gain net: {int(gain_net):,} FCFA"
        )
        await update.message.reply_text(msg)
    except Exception:
        await update.message.reply_text("Erreur. Format: /risque 1.30 2.21 | 1.45 3.10")
async def combo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text('Format: /combo "Barça vs Real" 1.30 2.21 | "Mali vs Sénégal" 1.45 3.10')
        return
    try:
        ligne = " ".join(context.args)
        mise_totale = MISE_DEFAUT
        tokens = ligne.split()
        try:
            derniere = float(tokens[-1])
            if derniere > 100 and '"' not in tokens[-1] and "|" not in tokens[-1]:
                mise_totale = derniere
                ligne = " ".join(tokens[:-1])
        except ValueError:
            pass
        segments = ligne.split("|")
        matchs, noms_matchs = [], []
        for seg in segments:
            seg = seg.strip()
            nom_match = re.search(r'"([^"]+)"', seg)
            if nom_match:
                nom = nom_match.group(1).strip()
                seg = seg.replace(nom_match.group(0), "").strip()
            else:
                nom = f"Match {len(matchs)+1}"
            cotes = [float(x) for x in seg.split() if x.strip()]
            if cotes:
                matchs.append(cotes)
                noms_matchs.append(nom)
        if len(matchs) < 2:
            await update.message.reply_text("Minimum 2 matchs separes par |")
            return
        if len(matchs) > 10:
            await update.message.reply_text("Maximum 10 matchs.")
            return
        resultats, nb_total, mise_par_combo = AlboraEngine.combo_matchs(matchs, noms_matchs, mise_totale)
        rentables = [r for r in resultats if r["rentable"]]
        gain_min = resultats[0]["gain_net"]
        gain_max = resultats[-1]["gain_net"]
        meilleure = rentables[0] if rentables else max(resultats, key=lambda x: x["gain_net"])
        lignes = []
        for i, r in enumerate(resultats, 1):
            signe = "+" if r["benefice_pct"] >= 0 else ""
            tag = "✅" if r["rentable"] else "❌"
            signe_b = "+" if r["benefice_net"] >= 0 else ""
            lignes.append(
                f"{tag} {i}. {r['noms_combo']}\n"
                f"   Mise: {r['mise_par_combo']:,} FCFA | Cote: {r['cote_totale']}\n"
                f"   Gain brut: {r['gain_brut']:,} FCFA | Taxe: -{r['taxe']:,} FCFA\n"
                f"   Gain net: {r['gain_net']:,} FCFA | Benefice: {signe_b}{r['benefice_net']:,} FCFA ({signe}{r['benefice_pct']}%) | Prob: {r['prob']}%"
            )
        signe_m = "+" if meilleure["benefice_pct"] >= 0 else ""
        meilleure_tag = "✅ RENTABLE" if meilleure["rentable"] else "NON RENTABLE"
        signe_bm = "+" if meilleure["benefice_net"] >= 0 else ""
        en_tete = (
            f"ALBORAA - {len(matchs)} matchs\n"
            f"Mise totale: {int(mise_totale):,} FCFA\n"
            f"Mise par combo: {mise_par_combo:,} FCFA\n"
            f"Seuil rentabilite: +{SEUIL_BENEFICE_FCFA:,} FCFA benefice net\n"
            f"Combos rentables: {len(rentables)}/{nb_total}\n"
            f"{'='*30}\n\n"
        )
        meilleure_section = (
            f"\n{'='*30}\n"
            f"MEILLEURE COMBO ({meilleure_tag})\n"
            f"{meilleure['noms_combo']}\n"
            f"Prob: {meilleure['prob']}% | Mise: {meilleure['mise_par_combo']:,} FCFA\n"
            f"Gain brut: {meilleure['gain_brut']:,} FCFA\n"
            f"Taxe (10%): -{meilleure['taxe']:,} FCFA\n"
            f"Gain net: {meilleure['gain_net']:,} FCFA\n"
            f"Benefice net: {signe_bm}{meilleure['benefice_net']:,} FCFA ({signe_m}{meilleure['benefice_pct']}%)\n"
        )
        avertissement = (
            f"\n{'='*30}\n"
            f"AVERTISSEMENT\n"
            f"Mise totale: {int(mise_totale):,} FCFA\n"
            f"Taxe 10% appliquee.\n"
            f"Une seule combo gagne.\n"
            f"Gain net min: {gain_min:,} FCFA\n"
            f"Gain net max: {gain_max:,} FCFA\n"
            f"Choisissez UNE seule combinaison."
        )
        _save_historique(update.effective_user.id, "combo", matchs, resultats)
        # Envoi par blocs de 10
        await update.message.reply_text(en_tete)
        for i in range(0, len(lignes), 10):
            bloc = "\n".join(lignes[i:i+10])
            if bloc.strip():
                await update.message.reply_text(bloc)
        await update.message.reply_text(meilleure_section + avertissement)
    except Exception:
        await update.message.reply_text('Erreur. Exemple: /combo "Barça vs Real" 1.30 2.21 | "Mali vs Sénégal" 1.45 3.10')
async def simuler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if len(args) != 2:
        await update.message.reply_text("Format: /simuler 1.85 10000")
        return
    try:
        cote, mise = float(args[0]), float(args[1])
        gain_brut = round(mise * cote, 0)
        taxe = round(gain_brut * TAXE, 0)
        gain_net = round(gain_brut - taxe, 0)
        benefice = round(gain_net - mise, 0)
        pct = round((gain_net / mise - 1) * 100, 2)
        msg = (
            f"SIMULATION\nCote: {cote}\nMise: {int(mise):,} FCFA\n"
            f"Gain brut: {int(gain_brut):,} FCFA\nTaxe (10%): -{int(taxe):,} FCFA\n"
            f"Gain net: {int(gain_net):,} FCFA\n"
            f"Benefice: {'+' if benefice >= 0 else ''}{int(benefice):,} FCFA ({'+' if pct >= 0 else ''}{pct}%)"
        )
        _save_historique(update.effective_user.id, "simuler", [cote, mise], {"gain": gain_net, "benefice": benefice})
        await update.message.reply_text(msg)
    except ValueError:
        await update.message.reply_text("Erreur. Exemple: /simuler 1.85 10000")
async def bankroll(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not context.args:
        await update.message.reply_text("Format: /bankroll 500000")
        return
    try:
        montant = float(context.args[0])
        if uid not in bankroll_data:
            bankroll_data[uid] = {"initial": montant, "actuel": montant, "total_mise": 0, "total_gain": 0, "nb_paris": 0}
            msg = f"BANKROLL INITIALISEE\nCapital: {int(montant):,} FCFA\nEnregistrer: /bankroll_pari 5000 15000"
        else:
            bankroll_data[uid]["actuel"] = montant
            msg = f"Bankroll mise a jour: {int(montant):,} FCFA"
        await update.message.reply_text(msg)
    except ValueError:
        await update.message.reply_text("Exemple: /bankroll 500000")
async def bankroll_pari(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if uid not in bankroll_data:
        await update.message.reply_text("Initialisez d'abord: /bankroll 500000")
        return
    if not context.args or len(context.args) != 2:
        await update.message.reply_text("Format: /bankroll_pari MISE GAIN_NET")
        return
    try:
        mise, gain = float(context.args[0]), float(context.args[1])
        profit = gain - mise
        b = bankroll_data[uid]
        b["total_mise"] += mise
        b["total_gain"] += gain
        b["actuel"] += profit
        b["nb_paris"] += 1
        roi = round((b["total_gain"] - b["total_mise"]) / b["total_mise"] * 100, 2)
        msg = (
            f"PARI ENREGISTRE\nMise: {int(mise):,} FCFA\nGain net: {int(gain):,} FCFA\n"
            f"Profit: {'+' if profit >= 0 else ''}{int(profit):,} FCFA\n\n"
            f"Bankroll: {int(b['actuel']):,} FCFA\nROI: {'+' if roi >= 0 else ''}{roi}%"
        )
        await update.message.reply_text(msg)
    except ValueError:
        await update.message.reply_text("Erreur. Exemple: /bankroll_pari 5000 15000")
async def bankroll_bilan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if uid not in bankroll_data:
        await update.message.reply_text("Aucune bankroll. Commencez: /bankroll 500000")
        return
    b = bankroll_data[uid]
    profit_total = b["actuel"] - b["initial"]
    roi = round(profit_total / b["initial"] * 100, 2) if b["initial"] > 0 else 0
    evolution = "📈" if profit_total >= 0 else "📉"
    msg = (
        f"BILAN BANKROLL\n{'='*30}\n"
        f"Capital initial: {int(b['initial']):,} FCFA\n"
        f"Capital actuel: {int(b['actuel']):,} FCFA\n"
        f"Evolution: {evolution} {'+' if profit_total >= 0 else ''}{int(profit_total):,} FCFA\n\n"
        f"Paris joues: {b['nb_paris']}\nTotal mise: {int(b['total_mise']):,} FCFA\n"
        f"Total gain: {int(b['total_gain']):,} FCFA\nROI: {'+' if roi >= 0 else ''}{roi}%"
    )
    await update.message.reply_text(msg)
async def objectif(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not context.args:
        await update.message.reply_text("Format: /objectif 1000000")
        return
    try:
        cible = float(context.args[0])
        capital_actuel = bankroll_data[uid]["actuel"] if uid in bankroll_data else MISE_DEFAUT
        objectif_data[uid] = {"cible": cible, "depart": capital_actuel}
        manque = cible - capital_actuel
        pct = round(capital_actuel / cible * 100, 1)
        msg = (
            f"OBJECTIF FIXE\nCible: {int(cible):,} FCFA\n"
            f"Capital actuel: {int(capital_actuel):,} FCFA\n"
            f"Manque: {int(manque):,} FCFA\nProgression: {pct}%"
        )
        await update.message.reply_text(msg)
    except ValueError:
        await update.message.reply_text("Exemple: /objectif 1000000")
async def objectif_bilan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if uid not in objectif_data:
        await update.message.reply_text("Aucun objectif. Fixez-en un: /objectif 1000000")
        return
    o = objectif_data[uid]
    capital_actuel = bankroll_data[uid]["actuel"] if uid in bankroll_data else o["depart"]
    manque = o["cible"] - capital_actuel
    pct = round(capital_actuel / o["cible"] * 100, 1)
    barre = int(pct / 10)
    barre_str = "█" * barre + "░" * (10 - barre)
    atteint = capital_actuel >= o["cible"]
    msg = (
        f"SUIVI OBJECTIF\n{'='*30}\n"
        f"Cible: {int(o['cible']):,} FCFA\n"
        f"Capital actuel: {int(capital_actuel):,} FCFA\n"
        f"Manque: {int(max(manque, 0)):,} FCFA\n\n"
        f"[{barre_str}] {pct}%\n\n"
        f"{'OBJECTIF ATTEINT!' if atteint else 'Continuez, vous progressez.'}"
    )
    await update.message.reply_text(msg)
async def bilan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if uid not in historique or not historique[uid]:
        await update.message.reply_text("Aucun historique pour cette session.")
        return
    lignes = [f"BILAN SESSION - {len(historique[uid])} operation(s)"]
    for i, h in enumerate(historique[uid], 1):
        lignes.append(f"{i}. [{h['type'].upper()}] {h['resume']} - {h['heure']}")
    await update.message.reply_text("\n".join(lignes))
async def effacer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    historique[update.effective_user.id] = []
    await update.message.reply_text("Historique efface.")
async def export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if uid not in historique or not historique[uid]:
        await update.message.reply_text("Aucune donnee a exporter.")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alboraa Session"
    headers = ["#", "Type", "Resume", "Heure"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
        cell.alignment = Alignment(horizontal="center")
    for i, h in enumerate(historique[uid], 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=h["type"].upper())
        ws.cell(row=i, column=3, value=h["resume"])
        ws.cell(row=i, column=4, value=h["heure"])
    ws.column_dimensions["C"].width = 60
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    await update.message.reply_document(
        document=buf,
        filename=f"alboraa_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        caption="Export session Alboraa"
    )
def _save_historique(uid, type_op, data, result):
    if uid not in historique:
        historique[uid] = []
    heure = datetime.now().strftime("%H:%M:%S")
    if type_op == "arbitrage":
        arb = result.get("arbitrage", False)
        resume = f"Cotes {data} -> {'ARB +' + str(result.get('profit','')) + '%' if arb else 'Pas arbitrage'}"
    elif type_op == "combo":
        rentables = [r for r in result if r.get("rentable")]
        resume = f"{len(data)} matchs -> {len(rentables)} combo(s) rentable(s)"
    elif type_op == "simuler":
        resume = f"Cote {data[0]} / Mise {int(data[1]):,} FCFA -> Gain net {int(result['gain']):,} FCFA"
    else:
        resume = str(data)
    historique[uid].append({"type": type_op, "resume": resume, "heure": heure})
# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    scanner_thread = threading.Thread(target=scanner_loop, args=(TELEGRAM_TOKEN,), daemon=True)
    scanner_thread.start()
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("alertes_on", alertes_on))
    app.add_handler(CommandHandler("alertes_off", alertes_off))
    app.add_handler(CommandHandler("scan", scan_manuel))
    app.add_handler(CommandHandler("sports", sports_cmd))
    app.add_handler(CommandHandler("exclure", exclure_bookmaker))
    app.add_handler(CommandHandler("inclure", inclure_bookmaker))
    app.add_handler(CommandHandler("liste_exclus", liste_exclus))
    app.add_handler(CommandHandler("arbitrage", arbitrage))
    app.add_handler(CommandHandler("combo", combo))
    app.add_handler(CommandHandler("value", value))
    app.add_handler(CommandHandler("risque", risque))
    app.add_handler(CommandHandler("simuler", simuler))
    app.add_handler(CommandHandler("bankroll", bankroll))
    app.add_handler(CommandHandler("bankroll_pari", bankroll_pari))
    app.add_handler(CommandHandler("bankroll_bilan", bankroll_bilan))
    app.add_handler(CommandHandler("objectif", objectif))
    app.add_handler(CommandHandler("objectif_bilan", objectif_bilan))
    app.add_handler(CommandHandler("bilan", bilan))
    app.add_handler(CommandHandler("effacer", effacer))
    app.add_handler(CommandHandler("export", export))
    print("Bot Alboraa demarre - version autonome v3")
    app.run_polling()
if __name__ == "__main__":
    main()
